import os
from datetime import datetime
from openpyxl import load_workbook

list_employee = [
    {'username1': '', 'username2' : '', 'full_name' : '', 'post' : ''},
    {'username1': '', 'username2' : '', 'full_name' : '', 'post' : ''},
]

class Act:
    def __init__(self):
        self.name_file_txt = 'data.txt'
        self.name_file_exl = 'example2.xlsx'
        self.final_data = {}
        self.current_user_data = self.user_data()

        self.list_month = {
            '01': 'января', '02': 'февраля', '03': 'марта',
            '04': 'апреля', '05': 'мая', '06': 'июня',
            '08': 'августа', '09': 'сентября', '07': 'июля',
            '10': 'октября', '11': 'ноября', '12': 'декабря',
        }

    def start(self):
        data = self.read_txt()
        self.creating_a_list(data)
        self.create_exl_file()

    def user_data(self):
        user = os.getlogin()
        for tmp in list_employee:
            if tmp['username1'] == user or tmp['username2'] == user:
                return tmp['full_name'], tmp['post']
        else:
            return None, None

    def read_txt(self):
        d = []

        if not os.path.exists(self.name_file_txt):
            print(f'File not found {self.name_file_txt}! \n Exiting the program')
            exit()

        with open(self.name_file_txt, 'r', encoding='utf-8') as file:
            for line in file.readlines():
                stripline = line.strip()
                if stripline and not stripline.startswith('#'):
                    d.append(stripline)
        return d

    def creating_a_list(self, data_list):
        if len(data_list) < 2:
            raise ValueError("Недостаточно данных: нужны ФИО и должность")

        self.final_data = {
            'user': {
                'name': data_list[0],
                'post': data_list[1]
            },
            'equipment': []
        }

        for i in range(2, len(data_list), 2):
            name = data_list[i]
            serial = data_list[i + 1] if i + 1 < len(data_list) else ''

            self.final_data['equipment'].append({
                'index': len(self.final_data['equipment']) + 1,
                'name': name,
                'serial': serial,
                'unit': 'шт',
                'quantity': 1
            })

    def create_exl_file(self):
        wb = load_workbook(self.name_file_exl)
        ws = wb['test']

        ws['R19'] = self.final_data['user']['name']
        ws['D19'] = self.final_data['user']['post']

        ws['R16'] = self.current_user_data[0]
        ws['D16'] = self.current_user_data[1]

        ws['E23'] = datetime.now().strftime("%d")
        ws['G23'] = self.list_month[datetime.now().strftime("%m")]
        ws['O23'] = datetime.now().strftime("%Y")

        # 2. Настраиваем параметры таблицы
        start_row = 6  # первая строка для техники
        num_items = len(self.final_data['equipment'])

        if num_items == 0:
            # Нет техники — ничего не делаем
            pass
        elif num_items == 1:
            # Одна строка — просто заполняем
            item = self.final_data['equipment'][0]
            ws[f'B{start_row}'] = item['index']
            ws[f'C{start_row}'] = item['name']
            ws[f'S{start_row}'] = item['serial']
            ws[f'X{start_row}'] = item['unit']
            ws[f'Y{start_row}'] = item['quantity']
        else:
            # Больше одной — вставляем (num_items - 1) строк после start_row
            # ws.insert_rows(start_row + 1, amount=num_items - 1)

            # Теперь заполняем все строки подряд
            for i, item in enumerate(self.final_data['equipment']):
                row = start_row + i
                ws[f'B{row}'] = item['index']
                ws[f'C{row}'] = item['name']
                ws[f'S{row}'] = item['serial']
                ws[f'X{row}'] = item['unit']
                ws[f'Y{row}'] = item['quantity']

        current_date = datetime.now().strftime("%d.%m.%Y")

        wb.save(f'{self.final_data["user"]["name"]}_{current_date}.xlsx')
        print('Done!')

if __name__== "__main__":
    act = Act()
    act.start()
